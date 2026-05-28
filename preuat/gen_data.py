import openpyxl

wb = openpyxl.load_workbook('c:/QA/CUTOVER/Tablero_Macroprocesos_N0-N2_v6.xlsx')
ws = wb['Trazabilidad (54 apps)']
rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if any(v is not None for v in row):
        r = [
            str(row[0]).strip().replace('\n',' ') if row[0] else None,
            str(row[1]).strip().replace('\n',' ') if row[1] else None,
            str(row[2]).strip().replace('\n',' ') if row[2] else None,
            str(row[3]).strip().replace('\n',' ') if row[3] else None,
            str(row[5]).strip().replace('\n',' ') if row[5] else None,
            str(row[6]).strip().replace('\n',' ') if row[6] else None,
            str(row[7]).strip().replace('\n',' ') if row[7] else None,
        ]
        rows.append(r)

lines = ['const RAW_DATA = [']
for r in rows:
    parts = []
    for v in r:
        if v is None:
            parts.append('null')
        else:
            escaped = v.replace('\\', '\\\\').replace('"', '\\"')
            parts.append(f'"{escaped}"')
    lines.append('[' + ','.join(parts) + '],')
lines.append('];')

with open('C:/QA/CUTOVER/raw_data_block.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Done: {len(rows)} rows written.")
